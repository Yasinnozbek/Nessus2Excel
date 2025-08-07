#!/usr/bin/env python3
import xml.etree.ElementTree as ET
import pandas as pd
import argparse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter

SEVERITY_COLOR_MAP = {
    4: "FF9999",   # Critical
    3: "FFCC99",   # High
    2: "FFFF99",   # Medium
    1: "CCFFCC",   # Low
}

def parse_nessus(file_path, output_file="nessus_grouped.xlsx"):
    tree = ET.parse(file_path)
    root = tree.getroot()
    report = root.find("Report")

    plugin_data = {}

    for host in report.findall("ReportHost"):
        host_ip = None
        for tag in host.findall("HostProperties/tag"):
            if tag.attrib.get("name") == "host-ip":
                host_ip = tag.text
                break

        for item in host.findall("ReportItem"):
            severity = int(item.attrib.get("severity", 0))
            if severity == 0:
                continue  # Skip severity 0

            plugin_id = item.attrib.get("pluginID")
            plugin_name = item.attrib.get("pluginName")
            port = item.attrib.get("port")
            protocol = item.attrib.get("protocol")
            svc_name = item.attrib.get("svc_name", protocol)
            description = item.findtext("description", "")
            solution = item.findtext("solution", "")
            cve = item.findtext("cve", "")
            output = item.findtext("plugin_output", "")
            see_alsos = [elem.text for elem in item.findall("see_also") if elem.text]
            refs = [elem.text for elem in item.findall("xref") if elem.text]

            key = plugin_id

            if key not in plugin_data:
                plugin_data[key] = {
                    "Plugin ID": plugin_id,
                    "Plugin Name": plugin_name,
                    "Severity": severity,
                    "CVE": cve,
                    "Description": description,
                    "Solution": solution,
                    "Protocol(s)": set(),
                    "Port(s)": set(),
                    "Affected IPs": [],
                    "Outputs": [],
                    "See Also": set(),
                    "Reference ID": set(),
                }

            plugin_data[key]["Protocol(s)"].add(protocol)
            plugin_data[key]["Port(s)"].add(port)

            ip_port = f"{host_ip} ({port}/{svc_name})"
            plugin_data[key]["Affected IPs"].append(ip_port)

            if output:
                plugin_data[key]["Outputs"].append(f"{host_ip} ({port}/{svc_name}):\n{output.strip()}")

            plugin_data[key]["See Also"].update(see_alsos)
            plugin_data[key]["Reference ID"].update(refs)

    # Build DataFrame
    rows = []
    for plugin in plugin_data.values():
        ips = sorted(plugin["Affected IPs"])
        ip_count = len(ips)  # Her IP+port ayrı sayılır
        output_text = "\n\n".join(plugin["Outputs"])
        see_also_text = "\n".join(sorted(plugin["See Also"]))
        ref_id_text = "\n".join(sorted(plugin["Reference ID"]))

        rows.append({
            "Plugin ID": plugin["Plugin ID"],
            "Plugin Name": plugin["Plugin Name"],
            "Severity": plugin["Severity"],
            "Affected IP Count": ip_count,
            "Affected IPs": "\n".join(ips),
            "Output": output_text,
            "CVE": plugin["CVE"],
            "Description": plugin["Description"],
            "Solution": plugin["Solution"],
            "Protocol(s)": ", ".join(sorted(plugin["Protocol(s)"])),
            "Port(s)": ", ".join(sorted(plugin["Port(s)"])),
            "See Also": see_also_text,
            "Reference ID": ref_id_text,
        })

    column_order = [
        "Plugin ID", "Plugin Name", "Severity",
        "Affected IP Count", "Affected IPs", "Output",
        "CVE", "Description", "Solution",
        "Protocol(s)", "Port(s)", "See Also", "Reference ID"
    ]

    df = pd.DataFrame(rows)[column_order]
    df.sort_values(by="Severity", ascending=False, inplace=True)

    df.to_excel(output_file, index=False)
    apply_formatting(output_file)
    print(f"✅ Exported to: {output_file}")

def apply_formatting(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 70)

    # Row coloring by severity
    severity_col_idx = None
    for i, cell in enumerate(ws[1], start=1):
        if cell.value == "Severity":
            severity_col_idx = i
            break

    if severity_col_idx:
        for row in ws.iter_rows(min_row=2):
            severity = int(row[severity_col_idx - 1].value)
            fill_color = SEVERITY_COLOR_MAP.get(severity)
            if fill_color:
                for cell in row:
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    wb.save(excel_file)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Convert .nessus to grouped Excel report.")
    parser.add_argument("file", help="Path to .nessus file")
    parser.add_argument("-o", "--output", default="nessus_grouped.xlsx", help="Output Excel file name")
    args = parser.parse_args()

    parse_nessus(args.file, args.output)
